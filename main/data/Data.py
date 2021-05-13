from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
from django.templatetags.static import static
from pathlib import Path
import os
import json

# Get main folder# Create product image directory if does not exist
# current_dir = os.getcwd()
# Path(f"{current_dir}/product image").mkdir(parents=True, exist_ok=True)
# Make product image path
# img_dir = f"{current_dir}/product image"

def openBook():
	# wb = load_workbook(static('data1.xlsx'))
	wb = load_workbook("C:\\Users\\CR\\Desktop\\OMS\\static\\data1.xlsx")
	ws1 = wb['Product Details']
	ws2 = wb['Order Form']
	return ws1, ws2

def makeJson(ws1, ws2):
	pic_name = 3
	image_loader = SheetImageLoader(ws1)
	title1 = []
	title2 = []
	datalist1 = []
	datalist2 = []
	data1 = {}
	data2 = {}

	# Get all from sheet1 
	for row in ws1.iter_rows(min_row=2, max_col=9, max_row=2):
		for cell in row:
			title1.append(cell.value)

	for row in range(3,139):
		for col in range(1,9):	
			letter = f'{get_column_letter(col)}{row}'
			if image_loader.image_in(letter):
				img_name = str(pic_name)+".png"
				# image = image_loader.get(letter)
				# img_path = f"{img_dir}/{pic_name}.png"
				# image.save(img_path)
				data1[f'{title1[col-1]}'] = img_name
			else:
				cell = ws1[letter]
				data1[f'{title1[col-1]}'] = cell.value
		pic_name+=1
		datalist1.append(data1)
		data1 = {}

	# Get info from sheet2
	for row in ws2.iter_rows(min_row=4, max_col=6, max_row=4):
		for cell in row:
			title2.append(cell.value)

	for row in range(5,141):
		for col in range(1,7):	
			if get_column_letter(col) == "A" or get_column_letter(col) == "F":
				letter = f'{get_column_letter(col)}{row}'
				cell = ws2[letter]
				data2[f'{title2[col-1]}'] = cell.value
		datalist2.append(data2)
		data2 = {}

	# Compare and put order multiple field into necessary objects
	for data2 in datalist2:
		for data1 in datalist1:
			if data1['SAP'] == data2['Sap Code']:
				data1['Order Multiples'] = data2['Order Multiples']


	# with open('product.json', 'w+') as output:
	# 	json.dump(datalist1, output, indent=4, separators=(',', ': '))
	# with open('form.json', 'w+') as output:
	# 	json.dump(datalist2, output, indent=4, separators=(',', ': '))
	# with open('data.json', 'w+') as output:
	# 	json.dump(datalist1, output, indent=4, separators=(',', ': '))
	# return datalist1

def makeData():
	ws1, ws2 = openBook()
	data = makeJson(ws1,ws2)
	return data

if __name__ == '__main__':
	makeData()


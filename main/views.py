from django.http import JsonResponse, Http404, HttpResponse, HttpResponseRedirect, HttpResponsePermanentRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from .decorators import checkLogin, checkSuperuser, checkCart, checkEdit
from django.templatetags.static import static
from django.utils import timezone
from django.core.mail import EmailMessage, BadHeaderError
from django.conf import settings
from django.contrib import messages
from .models import User, Product, Cart, Order, StoreSecret, WarehouseSecret, TrackingNumber
from .forms import RegisterForm
from .updateDB import updateData
import string
import json
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment

def index(request):
	# Currently Not needed
	# updateData()
	return redirect('logins')

@checkLogin
def register(request):
	if request.method == 'POST':
		form = RegisterForm(request.POST)
		if form.is_valid():
			user = form.save()
			messages.success(request, "Account Created")
			return redirect('logins')
		else:				
			errors = form.errors.get_json_data()
			for error in errors:
				message = errors[error][0]['message'] 
			messages.error(request, message)
	form = RegisterForm() 
	return render(request, 'main/register.html', {'form': form })

@checkLogin
def logins(request):
	nexts = ""
	if request.GET:  
		nexts = request.GET['next']
	if request.method == 'POST':
		form = AuthenticationForm(request, request.POST)
		if form.is_valid():
			username = request.POST['username']
			password = request.POST['password']
			user = authenticate(username=username, password=password)
			login(request, user)
			if nexts == "":
				return redirect("home")
			else:
				return HttpResponseRedirect(nexts)
		else:
			messages.error(request, "Incorrect Username or Password")
	form = AuthenticationForm()
	return render(request, 'main/login.html',context = {"form":form})

def logouts(request):
	logout(request)
	return redirect("index")

@login_required(login_url='logins')
def home(request):
	if request.user.allow_edit:
		orders = Order.objects.filter(checkout=True).order_by('-date')		
	else:
		orders = Order.objects.filter(user=request.user, checkout=True).order_by('-date')
	context = {'orders': orders}
	return render(request, 'main/home.html', context)

@login_required(login_url='logins')
def product(request):
	order, create = Order.objects.get_or_create(user=request.user, checkout=False)
	carts = order.cart_set.all()
	products = Product.objects.all()
	context = {'products':products, 'carts':carts}
	return render(request, 'main/product.html', context)

@login_required(login_url='logins')
def cart(request):
	order, create = Order.objects.get_or_create(user=request.user, checkout=False)
	carts = order.cart_set.all()
	context = {'carts':carts, 'order':order}
	return render(request, 'main/cart.html', context)

@login_required(login_url='logins')
@checkCart
def checkout(request):
	if request.method == "POST":
		confirmation_number = getConfirmationNumber()
		order = Order.objects.get(user=request.user, checkout=False)
		order.checkout = True
		order.comment = request.POST.get('comment')
		order.email = request.POST.get('email')
		order.confirmation_number = confirmation_number
		order.date = timezone.now()
		order.save()
		current_site = get_current_site(request)
		subject = "New Order"
		sender = settings.SENDER_EMAIL
		admin = settings.ADMIN_EMAIL
		warehouse = settings.WAREHOUSE_EMAIL
		store = request.POST.get('email')
		receiver = [store, admin, warehouse]
		carts = order.cart_set.all()
		content = render_to_string('main/checkoutemail.html', {
				'order': order,
				'carts': carts,
				'domain': current_site.domain,
				'user': request.user,
			})
		sendEmail(subject, content, sender, receiver)
		return HttpResponseRedirect(f'/ordered/{confirmation_number}')
	order, create = Order.objects.get_or_create(user=request.user, checkout=False)
	carts = order.cart_set.all()
	context = {'carts':carts, 'order':order}
	return render(request, 'main/checkout.html', context)

@login_required(login_url='logins')
def ordered(request, confirmation_number):
	return render(request, 'main/ordered.html', {'confirmation_number':confirmation_number})

@login_required(login_url='logins')
def viewOrder(request, confirmation_number):
	order = Order.objects.get(checkout=True, confirmation_number=confirmation_number)
	carts = order.cart_set.all()
	context = {'order':order, 'carts':carts}
	return render(request, 'main/vieworder.html', context)

@login_required(login_url='logins')
def updateCart(request):
	if request.is_ajax():
		product_id = request.POST.get('product_id')
		action = request.POST.get('action')
		product = Product.objects.get(id=product_id)
		order, create = Order.objects.get_or_create(user=request.user, checkout=False)
		cart, create = Cart.objects.get_or_create(order=order, product=product)
		if action == "add":
			cart.quantity += 1
			cart.save()
		elif action == "remove":
			cart.quantity -= 1
			cart.save()
		elif action == "delete":
			cart.quantity = 0
		cart.save()
		if cart.quantity <= 0:
			cart.delete()
		itemTotal = cart.getItemTotal
		itemQuantity = cart.getItemCount
		orderTotal = order.getCartTotal
		orderCount = order.getCartCount
		return JsonResponse({'action':action,'product_id':product_id,'quantity':cart.quantity,'itemTotal':itemTotal,'itemQuantity':itemQuantity, 'orderTotal':orderTotal,'orderCount':orderCount})
	return HttpResponseRedirect('/home')

def getConfirmationNumber():
	while True:
		confirmation_number = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k = 10))	
		if not Order.objects.filter(confirmation_number=confirmation_number).exists():
			break
	return confirmation_number

@login_required(login_url='logins')
@checkSuperuser
def generateSecretKey(request):
	while True:
		store_secret_key = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k = 10))	
		warehouse_secret_key = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k = 10))	
		if not StoreSecret.objects.filter(secret_key=store_secret_key).exists() and not WarehouseSecret.objects.filter(secret_key=warehouse_secret_key).exists():
			break
	if not StoreSecret.objects.filter(active=True).exists():
		storesecret = StoreSecret.objects.create(secret_key=store_secret_key, active=True)
	if not WarehouseSecret.objects.filter(active=True).exists():
		warehousesecret = WarehouseSecret.objects.create(secret_key=warehouse_secret_key, active=True)
	store_secret = StoreSecret.objects.get(active=True)
	warehouse_secret = WarehouseSecret.objects.get(active=True)
	context = {'store_secret_key':store_secret.secret_key, 'warehouse_secret_key':warehouse_secret.secret_key}
	return render(request, 'main/secret.html', context)

@login_required(login_url='logins')
@checkSuperuser
def updateStoreSecretKey(request):
	if request.is_ajax():
		while True:
			secret_key = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k = 10))	
			if not StoreSecret.objects.filter(secret_key=secret_key).exists():
				break
		secret = StoreSecret.objects.get(active=True)
		secret.delete()
		secret = StoreSecret.objects.create(secret_key=secret_key, active=True)
		context = {'secret_key':secret.secret_key}
		return JsonResponse(context)
	else:
		return HttpResponseRedirect('/home')

@login_required(login_url='logins')
@checkSuperuser
def updateWarehouseSecretKey(request):
	if request.is_ajax():
		while True:
			secret_key = ''.join(random.choices(string.ascii_uppercase + string.ascii_lowercase + string.digits, k = 10))	
			if not WarehouseSecret.objects.filter(secret_key=secret_key).exists():
				break
		secret = WarehouseSecret.objects.get(active=True)
		secret.delete()
		secret = WarehouseSecret.objects.create(secret_key=secret_key, active=True)
		context = {'secret_key':secret.secret_key}
		return JsonResponse(context)
	else:
		return HttpResponseRedirect('/home')

@login_required(login_url='logins')
@checkEdit
def editOrder(request, confirmation_number):
	if request.method == "POST":
		tracking_number = request.POST.get('tracking')
		order = Order.objects.get(checkout=True, confirmation_number=confirmation_number)
		carts = order.cart_set.all()
		for cart in carts:
			shippable = request.POST.get(f'{cart.product.sap}-shipped-quantity')
			if shippable != "":
				cart.shipped_quantity = int(shippable)
				cart.save()
		if tracking_number == "delete":
			tracking.delete()
		elif tracking_number != "" and order.getShippedCount != 0:
			tracking, create = TrackingNumber.objects.get_or_create(order=order, tracking_number=tracking_number)
			sender = settings.SENDER_EMAIL
			current_site = get_current_site(request)
			# admin = settings.ADMIN_EMAIL
			# warehouse = settings.WAREHOUSE_EMAIL
			store = order.email
			receiver = [store]
			subject = "Order Shipped"
			content = render_to_string('main/shippedemail.html', {
				'order': order,
				'carts': carts,
				'user': request.user,
				'domain': current_site.domain,

			})
			sendEmail(subject, content, sender, receiver)
		elif tracking_number != "" and order.getShippedCount == 0:
			messages.error(request, "Please Update Quantity Before Updating Tracking Info")
		if TrackingNumber.objects.filter(order=order).exists():
			order.confirm = True
			order.save()
		elif not TrackingNumber.objects.filter(order=order).exists():
			order.confirm = False
			order.save()
	order = Order.objects.get(checkout=True, confirmation_number=confirmation_number)
	carts = order.cart_set.all()
	tracking = order.trackingnumber_set.all()
	context = {'order':order, 'carts':carts, 'tracking':tracking}
	return render(request, 'main/editorder.html', context)

def sendEmail(subject, content, sender, receiver):
	email = EmailMessage(subject, content, sender, receiver)
	try: 
		email.send()
	except BadHeaderError:
		return HttpResponse("Invalid Header Found")

def exportExcel(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="Orders.xlsx"'
	wb = Workbook()
	ws1 = wb.create_sheet(title="Orders")
	# orders = list(Order.objects.filter(checkout=True))
	orders = Order.objects.filter(checkout=True, confirm=True)
	field = ['Order Date', 'Order Number', 'Store', 'SAP', 'DESC', 'Ordered Quantity', 'Ordered Unit', 'Shipped Quantity', 'Shipped Unit', 'Tracking Number']
	for row in range(1,2):
		for col in range(1, len(field)+1):
			ws1.cell(column=col, row=row, value=f"{field[col-1]}")
	order_number = 1
	row_start = 2
	row_end = row_start+1
	for order in orders:
		carts = order.cart_set.all()
		for cart in carts:
			for row in range(row_start, row_end):
				for col in range(1, len(field)+1):
					if col == 1:
						ws1.cell(column=col, row=row, value=f"{order.date.date()}")
					elif col == 2:
						ws1.cell(column=col, row=row, value=f"{order_number}")
					elif col == 3:
						ws1.cell(column=col, row=row, value=f"{order.user.store_name}-{order.user.store_location}")
					elif col == 4:
						ws1.cell(column=col, row=row, value=f"{cart.product.sap}")
					elif col == 5:
						ws1.cell(column=col, row=row, value=f"{cart.product.description}")
					elif col == 6:
						ws1.cell(column=col, row=row, value=f"{cart.quantity}")
					elif col == 7:
						ws1.cell(column=col, row=row, value=f"{cart.getItemCount}")
					elif col == 8:
						ws1.cell(column=col, row=row, value=f"{cart.shipped_quantity}")
					elif col == 9:
						ws1.cell(column=col, row=row, value=f"{cart.getShipableCount}")
					elif col == 10:
						track = ""
						for tracking in order.getTracking:
							track += f"{tracking}\n"
						ws1.cell(column=col, row=row).alignment = Alignment(wrapText=True)
						ws1.cell(column=col, row=row, value=f"{track}")
			row_start+=1
			row_end+=1
		order_number+=1
	# for row in range(2,len(orders)+2):
	# 	for col in range(1, len(field)+1):
	# 		order = orders[row-2]
	# 		carts = order.cart_set.all()
	# 		if col == 1:
	# 			ws1.cell(column=col, row=row, value=f"{order.date}")
	# 		elif col == 2:
	# 			ws1.cell(column=col, row=row, value=f"{order.user.store_name}-{order.user.store_location}")
	# 		elif col == 3:
	# 			detail = ""
	# 			for cart in carts:
	# 				detail += f"SAP({cart.product.sap})-Desc({cart.product.description})-Quantity({cart.quantity})-Shipped Quantity({cart.shipped_quantity})\n"
	# 			ws1.cell(column=col, row=row).alignment = Alignment(wrapText=True)
	# 			ws1.cell(column=col, row=row, value=detail)
	# 		elif col == 4:
	# 			ws1.cell(column=col, row=row, value=f"{order.tracking_number}")
	wb.save(response)
	return response


from django.db import models
from django.utils import timezone
from django.contrib.auth.models import AbstractUser
from django.templatetags.static import static
from openpyxl import Workbook
from openpyxl.styles import Alignment

class User(AbstractUser):
	id = models.AutoField(primary_key=True)
	store_name = models.CharField(max_length=100)
	store_location = models.PositiveSmallIntegerField(null=True)
	allow_edit = models.BooleanField(default=False)

	def __str__(self):
		return f"{self.username}"

class ReportDue(models.Model):
	id = models.AutoField(primary_key=True)
	date = models.DateTimeField(default=timezone.now, blank=True, null=True)

class StoreSecret(models.Model):
	id = models.AutoField(primary_key=True)
	secret_key = models.CharField(max_length=10, blank=True, null=True)
	active = models.BooleanField(default=False)

class WarehouseSecret(models.Model):
	id = models.AutoField(primary_key=True)
	secret_key = models.CharField(max_length=10, blank=True, null=True)
	active = models.BooleanField(default=False)

class Product(models.Model):
	id = models.AutoField(primary_key=True)
	category = models.CharField(max_length=100, null=True, blank=True)
	sap = models.IntegerField(null=True, blank=True)
	description = models.CharField(max_length=100, null=True, blank=True)
	picture = models.ImageField(null=True, blank=True)
	size = models.CharField(max_length=100, null=True, blank=True)
	unit_price = models.DecimalField(max_digits=6, decimal_places=2, blank=True, null=True)
	unit = models.IntegerField(null=True, blank=True)
	availability = models.BooleanField(default=True)
	
	def __str__(self):
		return f"{self.description}"

	@property
	def getTotal(self):
		if self.unit != None and self.unit_price != None:
			total = self.unit * self.unit_price
		else:
			total = 0
		return total
	

	@property
	def imageURL(self):
		try:
			url = self.picture.url
		except:
			url = static("unavailable.jpg")
		return url

class Order(models.Model):
	id = models.AutoField(primary_key=True)
	user = models.ForeignKey(User, null=True, blank=True, on_delete = models.SET_NULL)
	email = models.EmailField(null=True, blank=True)
	date = models.DateTimeField(default=timezone.now)
	checkout = models.BooleanField(default=False)
	confirm = models.BooleanField(default=False)
	confirmation_number = models.CharField(max_length=30, blank=True, null=True)
	comment = models.CharField(max_length=300, blank=True, null=True)

	def __str__(self):
		return f"user:{self.user}--checkout:{self.checkout}"

	# Return tracking number array
	@property
	def getTracking(self):
		tracking = self.trackingnumber_set.all()
		array = [track.tracking_number for track in tracking]
		return array
	

	# Return unqiue item count
	@property
	def getCartList(self):
		carts = self.cart_set.all()
		lists = len([cart.product for cart in carts])
		return lists
	
	# Return count of all items
	@property
	def getCartCount(self):
		carts = self.cart_set.all()
		total = sum([cart.getItemCount for cart in carts])
		return total

	# Return total price in cart
	@property
	def getCartTotal(self):
		carts = self.cart_set.all()
		total = sum([cart.getItemTotal for cart in carts])
		return total

	# Return shipped count of item
	@property
	def getShippedCount(self):
		carts = self.cart_set.all()
		lists = len([cart.shipped_quantity for cart in carts if cart.shipped_quantity != 0])
		return lists
	
	# Return total cost of shipped order
	@property
	def getShippedTotal(self):
		carts = self.cart_set.all()
		total = sum([cart.getShipableTotal for cart in carts])
		return total	

class TrackingNumber(models.Model):
	id = models.AutoField(primary_key=True)
	order = models.ForeignKey(Order, null=True, blank=True,on_delete = models.SET_NULL)
	tracking_number = models.CharField(max_length=100, blank=True, null=True)

	def __str__(self):
		return self.tracking_number


class Cart(models.Model):
	id = models.AutoField(primary_key=True)
	order = models.ForeignKey(Order, null=True, blank=True,on_delete = models.SET_NULL)
	product = models.ForeignKey(Product, null=True, blank=True,on_delete = models.SET_NULL)
	quantity = models.IntegerField(default=0, blank=True, null=True)
	shipped_quantity = models.IntegerField(default=0, blank=True, null=True)

	def __str__(self):
		return f"product:{self.product}"

	# Return total cost of item
	@property
	def getItemTotal(self):
		total = (self.product.unit_price *self.product.unit) * self.quantity
		return total
	
	# Return total count of item itself 
	@property
	def getItemCount(self):
		total = self.product.unit * self.quantity
		return total

	# Return total count of shipped item itself 
	@property
	def getShipableCount(self):
		total = self.product.unit * self.shipped_quantity
		return total

	# Return total cost of shipped item
	@property
	def getShipableTotal(self):
		total = (self.product.unit_price *self.product.unit) * self.shipped_quantity
		return total

def excel():
	wb = Workbook()
	ws1 = wb.create_sheet(title="Orders")
	# orders = list(Order.objects.filter(checkout=True))
	orders = Order.objects.filter(checkout=True, confirm=True)
	field = ['Order Date', 'Order Number', 'Store', 'SAP', 'DESC', 'Ordered Quantity', 'Ordered Unit', 'Shipped Quantity', 'Shipped Unit', 'Tracking Number']
	for row in range(1,2):
		for col in range(1, len(field)+1):
			ws1.cell(column=col, row=row, value=f"{field[col-1]}")
	order_number = 1
	row_start = 2
	row_end = row_start+1
	for order in orders:
		carts = order.cart_set.all()
		for cart in carts:
			for row in range(row_start, row_end):
				for col in range(1, len(field)+1):
					if col == 1:
						ws1.cell(column=col, row=row, value=f"{order.date.date()}")
					elif col == 2:
						ws1.cell(column=col, row=row, value=f"{order_number}")
					elif col == 3:
						ws1.cell(column=col, row=row, value=f"{order.user.store_name}-{order.user.store_location}")
					elif col == 4:
						ws1.cell(column=col, row=row, value=f"{cart.product.sap}")
					elif col == 5:
						ws1.cell(column=col, row=row, value=f"{cart.product.description}")
					elif col == 6:
						ws1.cell(column=col, row=row, value=f"{cart.quantity}")
					elif col == 7:
						ws1.cell(column=col, row=row, value=f"{cart.getItemCount}")
					elif col == 8:
						ws1.cell(column=col, row=row, value=f"{cart.shipped_quantity}")
					elif col == 9:
						ws1.cell(column=col, row=row, value=f"{cart.getShipableCount}")
					elif col == 10:
						track = ""
						for tracking in order.getTracking:
							track += f"{tracking}\n"
						ws1.cell(column=col, row=row).alignment = Alignment(wrapText=True)
						ws1.cell(column=col, row=row, value=f"{track}")
			row_start+=1
			row_end+=1
		order_number+=1
	return wb